#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
xlsx → CSV 独立转换工具

将 Excels 目录下所有 xlsx 文件的每个 Sheet 导出为独立的 CSV 文件。
不依赖 Luban schema，纯数据导出，适合数据查看和第三方系统导入。

用法:
    python xlsx2csv.py                          # 默认：Excels/ → Excels/CSV/
    python xlsx2csv.py --input ./Excels --output ./MyCSV
    python xlsx2csv.py --subdirs Tables,Local   # 只处理指定子目录
    python xlsx2csv.py --replace                # 转换后用 CSV 替换原 xlsx，删除临时目录
"""

import argparse
import csv
import os
import sys
from pathlib import Path


def check_openpyxl():
    """检查 openpyxl 是否安装，未安装则给出提示。"""
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("错误: 缺少 openpyxl 库。请执行以下命令安装：")
        print("  pip install openpyxl")
        sys.exit(1)


def ensure_dir(path: Path):
    """确保目录存在。"""
    path.mkdir(parents=True, exist_ok=True)


def is_row_empty(row: list) -> bool:
    """判断一行是否完全为空（所有单元格都是 None 或空字符串）。"""
    return all(cell is None or (isinstance(cell, str) and cell.strip() == "") for cell in row)


def cell_to_str(value) -> str:
    """将单元格值转为字符串。None → 空字符串，数字保留原样。"""
    if value is None:
        return ""
    if isinstance(value, float):
        # 整数浮点数去掉小数点（Excel 中 1 存储为 1.0）
        if value == int(value):
            return str(int(value))
    return str(value)


def convert_sheet_to_csv(ws, csv_path: Path) -> int:
    """
    将一个 Worksheet 导出为 CSV 文件。
    返回写入的数据行数（不含表头）。
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        data_row_count = 0
        header_written = False

        for row in rows:
            # 跳过完全为空的行
            if not header_written:
                writer.writerow([cell_to_str(c) for c in row])
                header_written = True
            elif is_row_empty(row):
                continue
            else:
                writer.writerow([cell_to_str(c) for c in row])
                data_row_count += 1

    return data_row_count


def safe_sheet_name(name: str) -> str:
    """清理 Sheet 名，替换文件名不安全的字符。"""
    forbidden = '<>:"/\\|?*'
    for ch in forbidden:
        name = name.replace(ch, "_")
    return name


def process_xlsx(xlsx_path: Path, output_dir: Path, relative_parent: Path) -> tuple:
    """
    处理单个 xlsx 文件，所有 Sheet 导出到 output_dir。
    返回 (成功Sheet数, 失败Sheet数)。
    """
    from openpyxl import load_workbook

    xlsx_stem = xlsx_path.stem  # 不含扩展名的文件名
    ok, fail = 0, 0

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  [错误] 无法打开 {xlsx_path.name}: {e}")
        return 0, 1

    multi_sheet = len(wb.sheetnames) > 1

    for sheet_name in wb.sheetnames:
        safe_name = safe_sheet_name(sheet_name)
        if multi_sheet:
            csv_name = f"{xlsx_stem}_{safe_name}.csv"
        else:
            csv_name = f"{xlsx_stem}.csv"
        csv_path = output_dir / csv_name

        try:
            ws = wb[sheet_name]
            row_count = convert_sheet_to_csv(ws, csv_path)
            print(f"  [OK] {csv_name} ({row_count} 行数据)")
            ok += 1
        except Exception as e:
            print(f"  [错误] 导出 Sheet '{sheet_name}' 失败: {e}")
            fail += 1

    wb.close()
    return ok, fail


def scan_xlsx_files(input_dir: Path, subdirs: list | None, output_dir_name: str) -> list:
    """
    扫描输入目录下的 xlsx 文件。
    返回 [(xlsx_path, output_dir), ...] 列表。
    """
    if subdirs is None:
        # 扫描所有子目录
        subdirs = [
            d.name for d in input_dir.iterdir()
            if d.is_dir() and d.name != output_dir_name and not d.name.startswith(".")
        ]

    tasks = []

    # 扫描子目录
    for sub in subdirs:
        sub_path = input_dir / sub
        if not sub_path.is_dir():
            print(f"  [警告] 目录不存在，跳过: {sub_path}")
            continue

        xlsx_files = sorted(sub_path.glob("*.xlsx"))
        if not xlsx_files:
            print(f"  [信息] {sub}/ 下无 xlsx 文件，跳过")
            continue

        output_dir = input_dir / output_dir_name / sub
        ensure_dir(output_dir)

        for xf in xlsx_files:
            tasks.append((xf, output_dir))

    # 扫描根目录下的 xlsx 文件
    root_xlsx = sorted(input_dir.glob("*.xlsx"))
    if root_xlsx:
        root_output = input_dir / output_dir_name
        ensure_dir(root_output)
        for xf in root_xlsx:
            tasks.append((xf, root_output))

    return tasks


def setup_console_encoding():
    """确保控制台输出使用 UTF-8（Windows 兼容）。"""
    if sys.platform == "win32":
        try:
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass


def main():
    setup_console_encoding()
    check_openpyxl()

    script_dir = Path(__file__).resolve().parent.parent  # Tools/
    default_input = script_dir.parent / "Config" / "Excels"

    parser = argparse.ArgumentParser(
        description="将 xlsx 配置表的每个 Sheet 导出为 CSV 文件"
    )
    parser.add_argument(
        "--input", "-i",
        type=Path,
        default=default_input,
        help=f"输入目录 (默认: {default_input})",
    )
    parser.add_argument(
        "--output", "-o",
        type=str,
        default="CSV",
        help="输出目录名，位于输入目录下 (默认: CSV)",
    )
    parser.add_argument(
        "--subdirs", "-s",
        type=str,
        default=None,
        help="要处理的子目录，逗号分隔 (默认: 输入目录下所有子目录)",
    )
    parser.add_argument(
        "--replace", "-r",
        action="store_true",
        help="用生成的 CSV 替换原 xlsx 文件，并删除临时输出目录",
    )
    args = parser.parse_args()

    input_dir: Path = args.input.resolve()
    output_dir_name: str = args.output
    subdirs: list | None = (
        [s.strip() for s in args.subdirs.split(",") if s.strip()]
        if args.subdirs
        else None
    )

    if not input_dir.is_dir():
        print(f"错误: 输入目录不存在: {input_dir}")
        sys.exit(1)

    print(f"输入目录: {input_dir}")
    print(f"输出目录: {input_dir / output_dir_name}")
    print(f"子目录:   {subdirs or '(自动扫描)'}")
    print("-" * 50)

    tasks = scan_xlsx_files(input_dir, subdirs, output_dir_name)
    if not tasks:
        print("未找到任何 xlsx 文件，退出。")
        return

    total_ok, total_fail, total_files = 0, 0, 0
    current_subdir = None
    converted_xlsx = []  # 记录成功转换的 xlsx 路径

    for xlsx_path, output_dir in tasks:
        # 根目录文件用 "(根目录)" 标识
        sub = xlsx_path.parent.name if xlsx_path.parent != input_dir else "(根目录)"
        if sub != current_subdir:
            current_subdir = sub
            print(f"\n[{sub}/]")

        print(f"  处理: {xlsx_path.name}")
        ok, fail = process_xlsx(xlsx_path, output_dir, xlsx_path.parent)
        total_ok += ok
        total_fail += fail
        total_files += 1
        if fail == 0:
            converted_xlsx.append(xlsx_path)

    print("\n" + "=" * 50)
    print(f"完成! 处理 {total_files} 个 xlsx 文件, {total_ok} 个 Sheet 成功, {total_fail} 个失败.")

    # --replace: 用 CSV 替换原 xlsx
    if args.replace and converted_xlsx:
        if total_fail > 0:
            print("[警告] 存在失败的 Sheet，仅替换完全成功转换的 xlsx 文件。")
        print("\n替换 xlsx → CSV ...")
        replaced = 0
        for xlsx_path in converted_xlsx:
            # 构造对应的 CSV 文件列表
            xlsx_stem = xlsx_path.stem
            is_root = xlsx_path.parent == input_dir
            csv_dir = input_dir / args.output
            if not is_root:
                csv_dir = csv_dir / xlsx_path.parent.name

            # 匹配该 xlsx 生成的所有 CSV
            csv_files = sorted(csv_dir.glob(f"{xlsx_stem}*.csv"))
            for cf in csv_files:
                dest = xlsx_path.parent / cf.name
                cf.replace(dest)
                print(f"  {cf.name} → {xlsx_path.parent.name + '/' if not is_root else ''}{cf.name}")

            # 删除原 xlsx
            xlsx_path.unlink()
            replaced += 1

        # 删除临时输出目录
        output_root = input_dir / args.output
        if output_root.exists():
            import shutil
            shutil.rmtree(output_root)
            print(f"\n已删除临时目录: {output_root}")

        print(f"替换完成! {replaced} 个 xlsx → CSV.")


if __name__ == "__main__":
    main()
